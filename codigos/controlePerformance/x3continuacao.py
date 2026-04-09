import time
import subprocess
from datetime import datetime, timedelta, date
import pyautogui
import pygetwindow as gw
import os
from openpyxl import load_workbook

WINDOW_TITLE_PART = "SISACOB"
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.05

def ativar_janela_excel_por_titulo(parte_nome):
    """Ativa a janela do Excel cujo título contém parte_nome."""
    janelas = [w for w in gw.getAllTitles() if parte_nome.lower() in w.lower()]
    if not janelas:
        return False
    janela = gw.getWindowsWithTitle(janelas[0])[0]
    if janela.isMinimized:
        janela.restore()
    janela.activate()
    try:
        janela.maximize()
    except Exception:
        pass
    time.sleep(0.7)
    return True

def janelas_excel():
    """Retorna set com títulos robustos de janelas do Excel abertas."""
    todas = gw.getAllTitles()
    return {t for t in todas if t and ("excel" in t.lower() or ".xls" in t.lower())}

def aguardar_janela_excel(antes, timeout=60):
    inicio = time.time()
    while time.time() - inicio < timeout:
        agora = janelas_excel()
        novas = agora - antes
        if novas:
            for titulo in novas:
                janelas = gw.getWindowsWithTitle(titulo)
                if janelas:
                    j = janelas[0]
                    if j.isMinimized:
                        j.restore()
                    j.activate()
                    try:
                        j.maximize()
                    except Exception:
                        pass
                    time.sleep(1)
                    return titulo
        time.sleep(0.5)
    return None

def focar_janela():
    janelas = gw.getWindowsWithTitle(WINDOW_TITLE_PART)
    if not janelas:
        return None
    janela = janelas[0]
    # Não maximiza nem minimiza — o SISACOB tem bugs com essas operações
    # Apenas traz para frente usando win32 se disponível, senão usa activate()
    try:
        import ctypes
        ctypes.windll.user32.SetForegroundWindow(janela._hWnd)
    except Exception:
        janela.activate()
    time.sleep(0.5)
    return janela

def main():
    janela = focar_janela()
    if not janela:
        return

    # Clica em "Monitoria"
    x = janela.left + 269
    y = janela.top + 40
    pyautogui.click(x, y)
    time.sleep(0.1)

    # Clica em "BSC Prioridade"
    x = janela.left + 368
    y = janela.top + 641
    pyautogui.click(x, y)
    time.sleep(0.5)

    # Garante o foco no campo de data antes de digitar
    x_data = janela.left + 90
    y_data = janela.top + 60
    pyautogui.click(x_data, y_data)
    time.sleep(0.1)

    # Digita o último dia útil anterior (formato ddmmaaaa) e tab
    data = datetime.now() - timedelta(days=1)
    # Retrocede até encontrar um dia útil (segunda a sexta), desconsiderando fim de semana
    while data.weekday() >= 5:  # 5 = sábado, 6 = domingo
        data -= timedelta(days=1)
    data_util = data.strftime("%d%m%Y")
    pyautogui.write(data_util, interval=0.03)
    pyautogui.press("tab")
    time.sleep(0.2)

    # Click relátorio de prod diaria
    x = janela.left + 203
    y = janela.top + 346
    pyautogui.click(x, y)
    time.sleep(0.5)

    # Registra janelas do Excel ANTES de gerar
    excels_antes = janelas_excel()

    # Click em gerar relatorio
    pyautogui.click(172, 444)
    time.sleep(1.7) #TEMPO ADICIONADO GERAR RELATORIO 02/04

    # Aguarda a planilha ser gerada sem mudar o foco para ela
    inicio_espera = time.time()
    while time.time() - inicio_espera < 30:
        novas = janelas_excel() - excels_antes
        if novas:
            break
        time.sleep(0.5)
    # Abre o arquivo REPORT_DIARIO_BUTTOW
    pasta_destino = r"X:\BACKOFFICE\BANCO ITAU\RELATORIOS\PERFORMANCE\2026\04. ABRIL\PRODUÇÃO DIÁRIA"
    arquivo_excel = os.path.join(pasta_destino, "REPORT_DIARIO_BUTTOW_0426_v4.xlsx")
    if not os.path.exists(arquivo_excel):
        return
    os.startfile(arquivo_excel)
    time.sleep(1.2)

    # Clique absoluto em (1168, 80) após digitar a data
    pyautogui.click(1168, 80)
    time.sleep(0.5)
    pyautogui.click(1168, 80)
    
    #click dia util
    time.sleep(0.8)
    pyautogui.click(218, 303)
    pyautogui.press('backspace')

    # Digita o último dia útil anterior ao dia atual
    today = date.today()
    # Calcula o número do dia útil do mês de hoje
    business_days = [d for d in range(1, today.day + 1) if date(today.year, today.month, d).weekday() < 5]
    numero_dia_util_hoje = len(business_days)
    # O dia útil anterior
    numero_dia_util_anterior = numero_dia_util_hoje - 1 if numero_dia_util_hoje > 1 else 1
    pyautogui.write(str(numero_dia_util_anterior), interval=0.05)
    time.sleep(0.5)
    pyautogui.click(397, 821)

    # Cópia e colagem entre planilhas
    copiar_bsc_prioridade_producaodia()

    # Volta ao SISACOB para nova consulta com intervalo mensal
    voltar_sisa_nova_consulta()

    arquivos = os.listdir(pasta_destino)
    bsc_marco = next((f for f in arquivos if 'abril' in f.lower() and 'ttow' in f.lower()), None)
    if bsc_marco:
        excels_antes = janelas_excel()
        os.startfile(os.path.join(pasta_destino, bsc_marco))
        # Aguarda a janela abrir e habilita edição
        aguardar_janela_excel(excels_antes, timeout=30)
        time.sleep(2)
        pyautogui.click(1163, 78)
        time.sleep(7)
        pyautogui.click(521, 173)
        # Atualiza todos os dados da planilha (clique em Dados > Atualizar Tudo)
        time.sleep(3)
        pyautogui.click(448, 45)
        time.sleep(3)
        pyautogui.click(456, 83)
        time.sleep(3)
        pyautogui.click(400, 826)
        time.sleep(3)
        pyautogui.click(448, 45)
        time.sleep(2)
        pyautogui.click(456, 83)

        # Clica na aba BÜTTOW
        time.sleep(1)
        pyautogui.click(244, 827)

        # Copia C42:F53 da aba BÜTTOW
        time.sleep(6)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(1)
        pyautogui.write('C42:F53')
        pyautogui.press('enter')
        time.sleep(0)
        pyautogui.hotkey('ctrl', 'c')

        # Volta para o REPORT_DIARIO_BUTTOW
        ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4')

        # Cola em F6 da aba BSC TOTAL
        subprocess.run(['powershell', '-command', "Set-Clipboard -Value \"'BSC TOTAL'!F6\""], check=False)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.1)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.3)
        pyautogui.press('enter')
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.2)
        pyautogui.press('ctrl')
        time.sleep(0.1)
        pyautogui.press('v')

        # Volta para a aba Início e salva
        time.sleep(0.3)
        pyautogui.hotkey('alt', 'h')
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 's')
    else:
        pass

    # Vai para aba Produção e deleta D5:U107
    ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4')
    time.sleep(0.5)
    subprocess.run(['powershell', '-command', 'Set-Clipboard -Value "Produção!D5:U107"'], check=False)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.3)
    pyautogui.press('enter')
    time.sleep(0.8)
    pyautogui.press('delete')

    # Abre o arquivo Performance Resultado - Abril
    pasta_abril = r"X:\BACKOFFICE\BANCO ITAU\RELATORIOS\PERFORMANCE\2026\04. ABRIL"
    arquivos_abril = os.listdir(pasta_abril)
    perf_arquivo = next((f for f in arquivos_abril if 'performance resultado' in f.lower() and 'abril' in f.lower() and 'backup' not in f.lower()), None)
    if perf_arquivo:
        caminho_perf = os.path.join(pasta_abril, perf_arquivo)
        excels_antes = janelas_excel()
        os.startfile(caminho_perf)
        aguardar_janela_excel(excels_antes, timeout=30)

        # Digita cpd@2025 no Performance Resultado - Abril
        time.sleep(0.5)
        pyautogui.click(1168, 80)
        time.sleep(0.5)
        pyautogui.click(1168, 80)
        time.sleep(0.3)
        subprocess.run(['powershell', '-command', 'Set-Clipboard -Value "cpd@2025"'], check=False)
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.2)
        pyautogui.press('enter')
        time.sleep(0.2)
        pyautogui.press('enter')

        # Copia B5:S10 da aba PRODUÇÃO DIÁRIA do Performance Resultado - Abril
        time.sleep(1)
        subprocess.run(['powershell', '-command', 'Set-Clipboard -Value "PRODUÇÃO DIÁRIA!B5:S10"'], check=False)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.1)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.3)
        pyautogui.press('enter')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')

        # Cola em D5 da aba Produção do REPORT_DIARIO_BUTTOW
        ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4')
        time.sleep(0.3)
        subprocess.run(['powershell', '-command', 'Set-Clipboard -Value "Produção!D5"'], check=False)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.1)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.3)
        pyautogui.press('enter')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'v')
    else:
        pass

def voltar_sisa_nova_consulta():
    janela = focar_janela()
    if not janela:
        return

    # Clica OK para fechar eventual diálogo
    pyautogui.click(871, 505)
    time.sleep(0.5)

    # Desmarca a caixinha
    pyautogui.click(194, 340)
    time.sleep(0.3)

    # Calcula primeiro dia do mês atual
    hoje = date.today()
    primeiro_dia = date(hoje.year, hoje.month, 1).strftime("%d%m%Y")

    # Calcula último dia útil do mês atual (último dia <= hoje sem ser fim de semana)
    ultimo_util = hoje
    while ultimo_util.weekday() >= 5:
        ultimo_util -= timedelta(days=1)
    ultimo_util_str = ultimo_util.strftime("%d%m%Y")

    # Clica no campo de data e preenche período
    pyautogui.click(108, 165)
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.write(primeiro_dia, interval=0.03)
    pyautogui.press('tab')
    time.sleep(0.15)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.write(ultimo_util_str, interval=0.03)
    pyautogui.press('tab')
    time.sleep(0.2)

    # Marca a nova opção
    pyautogui.click(195, 276)
    time.sleep(0.3)

    # Registra janelas do Excel ANTES de gerar
    excels_antes = janelas_excel()

    # Gerar relatório
    pyautogui.click(172, 444)
    time.sleep(1)

    # Aguarda nova janela do Excel
    excel = aguardar_janela_excel(excels_antes, timeout=30)

    # Clique final
    pyautogui.click(172, 444)
    time.sleep(0.5)

    # Copia C2:F2 da planilha BSCPRIORIDADECOLCHAO
    # Lê C2:F2 de BSCPRIORIDADECOLCHAO via openpyxl e cola direto em J26 do REPORT_DIARIO_BUTTOW
    pasta_destino = r"X:\BACKOFFICE\BANCO ITAU\RELATORIOS\PERFORMANCE\2026\04. ABRIL\PRODUÇÃO DIÁRIA"
    try:
        # Salva Performance Resultado - Abril antes de ler (garante dados atualizados no disco)
        if ativar_janela_excel_por_titulo('Performance Resultado - Abril'):
            pyautogui.hotkey('ctrl', 's')
            time.sleep(1.5)

        # Descobre o nome real do arquivo na pasta
        arquivos = os.listdir(pasta_destino)
        colchao_arquivo = next((f for f in arquivos if 'performance resultado' in f.lower() or 'performance' in f.lower() and 'abril' in f.lower()), None)
        if not colchao_arquivo:
            return
        caminho_colchao = os.path.join(pasta_destino, colchao_arquivo)

        wb_colchao = load_workbook(caminho_colchao, data_only=True)
        ws_colchao = wb_colchao.active
        valores = [ws_colchao.cell(row=2, column=c).value for c in range(3, 7)]  # C2:F2
        wb_colchao.close()

        # Fecha o REPORT_DIARIO_BUTTOW no Excel antes de salvar com openpyxl
        if ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4'):
            pyautogui.hotkey('ctrl', 'w')
            time.sleep(1)

        caminho_report = os.path.join(pasta_destino, 'REPORT_DIARIO_BUTTOW_0426_v4.xlsx')
        wb_report = load_workbook(caminho_report)
        ws_report = wb_report['BSC TOTAL']
        for i, valor in enumerate(valores):
            ws_report.cell(row=26, column=10 + i, value=valor)  # J=10
        wb_report.save(caminho_report)
        wb_report.close()

        # Reabre o REPORT_DIARIO_BUTTOW
        os.startfile(caminho_report)
        time.sleep(1.5)
    except Exception as e:
        pass

def copiar_bsc_prioridade_producaodia():
    # Ativa a janela da planilha de origem
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        return
    # Seleciona o intervalo C2:F13
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.2)
    pyautogui.write('C2:F13')
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'c')

    # Ativa a janela da planilha de destino
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4'):
        return
    # Seleciona a célula G6
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.2)
    pyautogui.write('G6')
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'v')

    # Volta para BSC PRIORIDADE e navega diretamente para aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        return
    time.sleep(0.3)
    subprocess.run(['powershell', '-command', 'Set-Clipboard -Value "PRECIFICACAO!C9"'], check=False)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.3)
    pyautogui.press('enter')
    time.sleep(0.2)

    # Após navegar para a aba, insere explicitamente as fórmulas em C9, D9, E9, F9
    for col in ['C', 'D', 'E', 'F']:
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.05)
        pyautogui.write(f'{col}9')
        pyautogui.press('enter')
        time.sleep(0.05)
        pyautogui.write(f'={col}7+{col}8', interval=0)
        pyautogui.press('enter')
        time.sleep(0.05)

    # Copia C2:F7 da aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        return
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.1)
    pyautogui.write('C2:F6')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')

    # Cola C2:F7 em G6 da aba BSC - PRECIF no REPORT_DIARIO_BUTTOW
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4'):
        return
    time.sleep(0.5)
    subprocess.run(['powershell', '-command', "Set-Clipboard -Value \"'BSC - PRECIF'!G6\""], check=False)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.3)
    pyautogui.press('enter')
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'alt', 'v')
    time.sleep(0.3)
    pyautogui.press('v')
    pyautogui.press('enter')

    # Copia C9:F9 da aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        return
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.1)
    pyautogui.write('C9:F9')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')

    # Cola C9:F9 em G13 da aba BSC - PRECIF no REPORT_DIARIO_BUTTOW (pulando linha 8)
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0426_v4'):
        return
    time.sleep(0.5)
    subprocess.run(['powershell', '-command', "Set-Clipboard -Value \"'BSC - PRECIF'!G11\""], check=False)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.3)
    pyautogui.press('enter')
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'alt', 'v')
    time.sleep(0.3)
    pyautogui.press('v')
    pyautogui.press('enter')

if __name__ == "__main__":
    main()