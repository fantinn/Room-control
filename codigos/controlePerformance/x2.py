import time
from datetime import datetime, timedelta, date
import pyautogui
import pygetwindow as gw
import os
from openpyxl import load_workbook

WINDOW_TITLE_PART = "SISACOB"
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.05

def ativar_janela_excel_por_titulo(parte_nome):
    """Ativa a janela do Excel cujo título contém parte_nome."""
    janelas = [w for w in gw.getAllTitles() if parte_nome.lower() in w.lower()]
    if not janelas:
        print(f'[ERRO] Nenhuma janela do Excel encontrada contendo: {parte_nome}')
        return False
    janela = gw.getWindowsWithTitle(janelas[0])[0]
    if janela.isMinimized:
        janela.restore()
    janela.activate()
    try:
        janela.maximize()
    except Exception:
        pass
    print(f'[INFO] Janela do Excel ativada: {janelas[0]}')
    time.sleep(0.7)
    return True

def janelas_excel():
    """Retorna set com títulos robustos de janelas do Excel abertas."""
    todas = gw.getAllTitles()
    return {t for t in todas if t and ("excel" in t.lower() or ".xls" in t.lower())}

def aguardar_janela_excel(antes, timeout=60):
    print("[INFO] Aguardando nova janela do Excel abrir...")
    inicio = time.time()
    while time.time() - inicio < timeout:
        agora = janelas_excel()
        novas = agora - antes
        if novas:
            for titulo in novas:
                print(f"[INFO] Nova janela do Excel detectada: {titulo}")
                janelas = gw.getWindowsWithTitle(titulo)
                if janelas:
                    j = janelas[0]
                    if j.isMinimized:
                        j.restore()
                    j.activate()
                    try:
                        j.maximize()
                    except Exception:
                        pass
                    print("[INFO] Excel focado e maximizado.")
                    time.sleep(1)
                    return titulo
        time.sleep(0.5)
    print("[ERRO] Nenhuma janela do Excel detectada no tempo limite.")
    return None

def focar_janela():
    print("[INFO] Procurando janela SISACOB...")
    janelas = gw.getWindowsWithTitle(WINDOW_TITLE_PART)
    if not janelas:
        print("[ERRO] Janela do SISACOB não encontrada. O sistema está aberto?")
        return None
    janela = janelas[0]
    # Não maximiza nem minimiza — o SISACOB tem bugs com essas operações
    # Apenas traz para frente usando win32 se disponível, senão usa activate()
    try:
        import ctypes
        ctypes.windll.user32.SetForegroundWindow(janela._hWnd)
    except Exception:
        janela.activate()
    print("[INFO] Janela SISACOB focada.")
    time.sleep(0.5)
    return janela

def main():
    janela = focar_janela()
    if not janela:
        print("[FIM] Não foi possível encontrar a janela SISACOB. Encerrando script.")
        return

    # Clica em "Monitoria"
    x = janela.left + 269
    y = janela.top + 40
    print(f"[INFO] Clicando em 'Monitoria' na posição ({x}, {y})...")
    pyautogui.click(x, y)
    time.sleep(0.1)

    # Clica em "BSC Prioridade"
    x = janela.left + 368
    y = janela.top + 641
    print(f"[INFO] Clicando em 'BSC Prioridade' na posição ({x}, {y})...")
    pyautogui.click(x, y)
    time.sleep(0.5)

    # Garante o foco no campo de data antes de digitar
    x_data = janela.left + 90
    y_data = janela.top + 60
    print(f"[INFO] Clicando no campo de data em ({x_data}, {y_data})...")
    pyautogui.click(x_data, y_data)
    time.sleep(0.1)

    # Digita o último dia útil anterior (formato ddmmaaaa) e tab
    data = datetime.now() - timedelta(days=1)
    # Retrocede até encontrar um dia útil (segunda a sexta), desconsiderando fim de semana
    while data.weekday() >= 5:  # 5 = sábado, 6 = domingo
        data -= timedelta(days=1)
    data_util = data.strftime("%d%m%Y")
    print(f"[INFO] Digitando data útil: {data_util}")
    pyautogui.write(data_util, interval=0.03)
    pyautogui.press("tab")
    time.sleep(0.2)

    # Click relátorio de prod diaria
    x = janela.left + 203
    y = janela.top + 346
    print(f"[INFO] Clicando em 'Relatório de Produção Diária' na posição ({x}, {y})...")
    pyautogui.click(x, y)
    time.sleep(0.5)

    # Registra janelas do Excel ANTES de gerar
    excels_antes = janelas_excel()

    # Click em gerar relatorio
    print("[INFO] Clicando em 'Gerar Relatório' em (172, 444)...")
    pyautogui.click(172, 444)
    time.sleep(1.7) #TEMPO ADICIONADO GERAR RELATORIO 02/04

    # Aguarda a planilha ser gerada sem mudar o foco para ela
    print("[INFO] Aguardando geração do relatório sem mudar de janela...")
    inicio_espera = time.time()
    while time.time() - inicio_espera < 30:
        novas = janelas_excel() - excels_antes
        if novas:
            print(f"[INFO] Relatório gerado detectado: {novas}")
            break
        time.sleep(0.5)
    else:
        print("[AVISO] Nenhuma nova janela do Excel detectada no tempo limite, continuando mesmo assim.")

    # Abre o arquivo REPORT_DIARIO_BUTTOW
    pasta_destino = r"X:\BACKOFFICE\BANCO ITAU\RELATORIOS\PERFORMANCE\2026\03. MARÇO\PRODUÇÃO DIÁRIA"
    arquivo_excel = os.path.join(pasta_destino, "REPORT_DIARIO_BUTTOW_0326_v4.xlsx")
    print(f"[INFO] Abrindo arquivo: {arquivo_excel}")
    if not os.path.exists(arquivo_excel):
        print(f"[ERRO] Arquivo não encontrado: {arquivo_excel}")
        return
    os.startfile(arquivo_excel)
    time.sleep(1.2)

    # Clique absoluto em (1168, 80) após digitar a data
    pyautogui.click(1168, 80)
    
    #click dia util
    time.sleep(0.8)
    pyautogui.click(218, 303)
    pyautogui.press('backspace')

    # Digita o último dia útil anterior ao dia atual
    today = date.today()
    # Calcula o número do dia útil do mês de hoje
    business_days = [d for d in range(1, today.day + 1) if date(today.year, today.month, d).weekday() < 5]
    numero_dia_util_hoje = len(business_days)
    # O dia útil anterior
    numero_dia_util_anterior = numero_dia_util_hoje - 1 if numero_dia_util_hoje > 1 else 1
    pyautogui.write(str(numero_dia_util_anterior), interval=0.05)
    time.sleep(0.5)
    pyautogui.click(397, 821)
    print('[INFO] Clique final realizado em (397, 821)')

    # Cópia e colagem entre planilhas
    copiar_bsc_prioridade_producaodia()

    # Volta ao SISACOB para nova consulta com intervalo mensal
    voltar_sisa_nova_consulta()

    arquivos = os.listdir(pasta_destino)
    bsc_marco = next((f for f in arquivos if 'b' in f.lower() and 'ttow' in f.lower() and 'report' not in f.lower()), None)
    if bsc_marco:
        excels_antes = janelas_excel()
        os.startfile(os.path.join(pasta_destino, bsc_marco))
        print(f'[INFO] Arquivo aberto: {bsc_marco}')
        # Aguarda a janela abrir e habilita edição
        aguardar_janela_excel(excels_antes, timeout=30)
        time.sleep(2)
        pyautogui.click(1163, 78)
        print('[INFO] Habilitar edição clicado em (1163, 78).')
        time.sleep(7)
        pyautogui.click(521, 173)
        print('[INFO] Clique em (521, 173) realizado.')
        # Atualiza todos os dados da planilha (clique em Dados > Atualizar Tudo)
        time.sleep(3)
        pyautogui.click(448, 45)
        print('[INFO] Clique em (448, 45) para Atualizar Tudo.')
        time.sleep(3)
        pyautogui.click(456, 83)
        print('[INFO] Clique em (456, 83) realizado.')
        time.sleep(3)
        pyautogui.click(400, 826)
        print('[INFO] Clique em (400, 826) realizado.')
        time.sleep(3)
        pyautogui.click(448, 45)
        print('[INFO] Clique em (448, 45) — segunda atualização.')
        time.sleep(2)
        pyautogui.click(456, 83)
        print('[INFO] Clique em (456, 83) — segunda atualização.')

        # Clica na aba BÜTTOW
        time.sleep(1)
        pyautogui.click(244, 827)
        print('[INFO] Clique em (244, 827) — aba BÜTTOW.')

        # Copia C42:F53 da aba BÜTTOW
        time.sleep(6)
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(1)
        pyautogui.write('C42:F53')
        pyautogui.press('enter')
        time.sleep(0)
        pyautogui.hotkey('ctrl', 'c')
        print('[INFO] C42:F53 copiado da aba BÜTTOW.')

        # Volta para o REPORT_DIARIO_BUTTOW
        ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0326_v4')

        # Cola em F6 da aba BSC TOTAL
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.2)
        pyautogui.write("'BSC TOTAL'!F6")
        pyautogui.press('enter')
        time.sleep(0.2)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.2)
        pyautogui.press('ctrl')
        time.sleep(0.1)
        pyautogui.press('v')
        print('[INFO] C42:F53 colado em F6 da aba BSC TOTAL.')
    else:
        print('[ERRO] Arquivo BSC MARÇO - BÜTTOW_2026 não encontrado!')

def voltar_sisa_nova_consulta():
    janela = focar_janela()
    if not janela:
        print('[ERRO] Não foi possível focar o SISACOB para nova consulta.')
        return

    # Clica OK para fechar eventual diálogo
    print('[INFO] Clicando OK em (871, 505)...')
    pyautogui.click(871, 505)
    time.sleep(0.5)

    # Desmarca a caixinha
    print('[INFO] Desmarcando checkbox em (194, 340)...')
    pyautogui.click(194, 340)
    time.sleep(0.3)

    # Calcula primeiro dia do mês atual
    hoje = date.today()
    primeiro_dia = date(hoje.year, hoje.month, 1).strftime("%d%m%Y")

    # Calcula último dia útil do mês atual (último dia <= hoje sem ser fim de semana)
    ultimo_util = hoje
    while ultimo_util.weekday() >= 5:
        ultimo_util -= timedelta(days=1)
    ultimo_util_str = ultimo_util.strftime("%d%m%Y")

    # Clica no campo de data e preenche período
    print(f'[INFO] Clicando no campo de data em (108, 165)...')
    pyautogui.click(108, 165)
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.write(primeiro_dia, interval=0.03)
    print(f'[INFO] Primeira data digitada: {primeiro_dia}')
    pyautogui.press('tab')
    time.sleep(0.15)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.write(ultimo_util_str, interval=0.03)
    print(f'[INFO] Última data útil digitada: {ultimo_util_str}')
    pyautogui.press('tab')
    time.sleep(0.2)

    # Marca a nova opção
    print('[INFO] Marcando checkbox em (195, 276)...')
    pyautogui.click(195, 276)
    time.sleep(0.3)

    # Registra janelas do Excel ANTES de gerar
    excels_antes = janelas_excel()

    # Gerar relatório
    print("[INFO] Clicando em 'Gerar Relatório' em (172, 444)...")
    pyautogui.click(172, 444)
    time.sleep(1)

    # Aguarda nova janela do Excel
    excel = aguardar_janela_excel(excels_antes, timeout=30)
    if not excel:
        print('[AVISO] Nenhuma janela do Excel detectada após gerar o segundo relatório.')

    # Clique final
    print('[INFO] Clicando em (172, 444)...')
    pyautogui.click(172, 444)
    time.sleep(0.5)

    # Copia C2:F2 da planilha BSCPRIORIDADECOLCHAO
    # Lê C2:F2 de BSCPRIORIDADECOLCHAO via openpyxl e cola direto em J26 do REPORT_DIARIO_BUTTOW
    pasta_destino = r"X:\BACKOFFICE\BANCO ITAU\RELATORIOS\PERFORMANCE\2026\03. MARÇO\PRODUÇÃO DIÁRIA"
    try:
        # Salva BSCPRIORIDADECOLCHAO antes de ler (garante dados atualizados no disco)
        if ativar_janela_excel_por_titulo('BSCPRIORIDADECOLCHAO'):
            pyautogui.hotkey('ctrl', 's')
            time.sleep(1.5)

        # Descobre o nome real do arquivo na pasta
        arquivos = os.listdir(pasta_destino)
        colchao_arquivo = next((f for f in arquivos if 'colchao' in f.lower() or 'colchão' in f.lower()), None)
        if not colchao_arquivo:
            print('[ERRO] Arquivo BSCPRIORIDADECOLCHAO não encontrado na pasta!')
            return
        caminho_colchao = os.path.join(pasta_destino, colchao_arquivo)
        print(f'[INFO] Lendo arquivo: {caminho_colchao}')

        wb_colchao = load_workbook(caminho_colchao, data_only=True)
        ws_colchao = wb_colchao.active
        valores = [ws_colchao.cell(row=2, column=c).value for c in range(3, 7)]  # C2:F2
        wb_colchao.close()
        print(f'[INFO] Valores lidos de C2:F2: {valores}')

        # Fecha o REPORT_DIARIO_BUTTOW no Excel antes de salvar com openpyxl
        if ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0326_v4'):
            pyautogui.hotkey('ctrl', 'w')
            time.sleep(1)

        caminho_report = os.path.join(pasta_destino, 'REPORT_DIARIO_BUTTOW_0326_v4.xlsx')
        wb_report = load_workbook(caminho_report)
        ws_report = wb_report['BSC TOTAL']
        for i, valor in enumerate(valores):
            ws_report.cell(row=26, column=10 + i, value=valor)  # J=10
        wb_report.save(caminho_report)
        wb_report.close()
        print('[INFO] C2:F2 colado em J26 da aba BSC TOTAL via openpyxl.')

        # Reabre o REPORT_DIARIO_BUTTOW
        os.startfile(caminho_report)
        time.sleep(1.5)
    except Exception as e:
        print(f'[ERRO] Falha ao copiar C2:F2 para J26: {e}')

def copiar_bsc_prioridade_producaodia():
    # Ativa a janela da planilha de origem
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        print('[ERRO] Não foi possível ativar a janela da planilha de prioridade!')
        return
    # Seleciona o intervalo C2:F13
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.2)
    pyautogui.write('C2:F13')
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'c')
    print('[INFO] Intervalo C2:F13 copiado da planilha de prioridade!')

    # Ativa a janela da planilha de destino
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0326_v4'):
        print('[ERRO] Não foi possível ativar a janela da planilha de destino!')
        return
    # Seleciona a célula G6
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.2)
    pyautogui.write('G6')
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'v')
    print('[INFO] Dados colados em G6 da planilha de destino!')

    # Volta para BSC PRIORIDADE e navega diretamente para aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        print('[ERRO] Não foi possível reativar a janela de prioridade!')
        return
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.15)
    pyautogui.write('PRECIFICACAO!C9')
    pyautogui.press('enter')
    time.sleep(0.2)
    print('[INFO] Navegado para aba PRECIFICACAO.')

    # Após navegar para a aba, insere explicitamente as fórmulas em C9, D9, E9, F9
    for col in ['C', 'D', 'E', 'F']:
        pyautogui.hotkey('ctrl', 'g')
        time.sleep(0.05)
        pyautogui.write(f'{col}9')
        pyautogui.press('enter')
        time.sleep(0.05)
        pyautogui.write(f'={col}7+{col}8', interval=0)
        pyautogui.press('enter')
        time.sleep(0.05)
        print(f'[INFO] Fórmula ={col}7+{col}8 inserida em {col}9.')

    # Copia C2:F7 da aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        print('[ERRO] Não foi possível garantir o foco na planilha de prioridade para copiar C2:F7!')
        return
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.1)
    pyautogui.write('C2:F6')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')
    print('[INFO] Intervalo C2:F6 copiado da aba PRECIFICACAO.')

    # Cola C2:F7 em G6 da aba BSC - PRECIF no REPORT_DIARIO_BUTTOW
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0326_v4'):
        print('[ERRO] Não foi possível ativar REPORT_DIARIO_BUTTOW para colar C2:F7!')
        return
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.15)
    pyautogui.write("'BSC - PRECIF'!G6")
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'alt', 'v')
    time.sleep(0.3)
    pyautogui.press('v')
    pyautogui.press('enter')
    print('[INFO] C2:F7 colado (apenas valores) em G6 da aba BSC - PRECIF.')

    # Copia C9:F9 da aba PRECIFICACAO
    if not ativar_janela_excel_por_titulo('BSC PRIORIDADE - PRODUCAODIA'):
        print('[ERRO] Não foi possível reativar a planilha de prioridade para copiar C9:F9!')
        return
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.1)
    pyautogui.write('C9:F9')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')
    print('[INFO] Intervalo C9:F9 copiado da aba PRECIFICACAO.')

    # Cola C9:F9 em G13 da aba BSC - PRECIF no REPORT_DIARIO_BUTTOW (pulando linha 8)
    if not ativar_janela_excel_por_titulo('REPORT_DIARIO_BUTTOW_0326_v4'):
        print('[ERRO] Não foi possível ativar REPORT_DIARIO_BUTTOW para colar C9:F9!')
        return
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'g')
    time.sleep(0.15)
    pyautogui.write("'BSC - PRECIF'!G11")
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'alt', 'v')
    time.sleep(0.3)
    pyautogui.press('v')
    pyautogui.press('enter')
    print('[INFO] C9:F9 colado (apenas valores) em G11 da aba BSC - PRECIF.')

if __name__ == "__main__":
    main()